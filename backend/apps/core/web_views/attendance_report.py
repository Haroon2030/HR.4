"""تقرير الحضور اليومي — تجميع بصمات الدخول والخروج."""
from datetime import datetime

from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.views.decorators.http import require_GET

from apps.attendance.models import AttendancePunch
from apps.attendance.selectors.daily_report import (
    build_daily_attendance_rows,
    daily_rows_to_table,
    summarize_daily_rows,
)
from apps.attendance.selectors.punch_export import punches_to_table_rows
from apps.attendance.selectors.punch_records import PUNCH_LIST_ORDERING, get_punch_queryset, get_punch_stats
from apps.attendance.selectors.biometric_devices import (
    filter_biometric_devices_for_user,
    get_biometric_devices_queryset,
)
from apps.core.decorators import permission_required
from apps.core.models import Branch
from apps.core.web_views._helpers import _user_accessible_branch_ids
from apps.core.web_views.attendance_records import (
    _apply_default_date_filters,
    _filters_to_querystring,
    _parse_filters,
)
from apps.employees.models import Employee


def _punches_for_report(request, filters: dict):
    date_from = None
    date_to = None
    if filters['date_from']:
        date_from = datetime.strptime(filters['date_from'], '%Y-%m-%d').date()
    if filters['date_to']:
        date_to = datetime.strptime(filters['date_to'], '%Y-%m-%d').date()

    qs = get_punch_queryset(
        device_id=filters['device_id'],
        branch_ids=filters['branch_ids'],
        employee_id=filters['employee_id'],
        device_user_id=filters['device_user_id'],
        date_from=date_from,
        date_to=date_to,
        punch_type=filters['punch_type'],
        mapped_only=filters['mapped_only'],
        search=filters['search'] or None,
    )
    return qs.filter(device_id__in=filter_biometric_devices_for_user(request.user).values('pk'))


@permission_required('attendance.view')
def attendance_report(request):
    filters = _apply_default_date_filters(_parse_filters(request))
    qs = _punches_for_report(request, filters)
    punch_stats = get_punch_stats(qs)
    all_rows = build_daily_attendance_rows(qs)
    summary = summarize_daily_rows(all_rows, punch_total=punch_stats['total'])

    daily_per_page = int(request.GET.get('per_page') or 50)
    daily_paginator = Paginator(all_rows, per_page=daily_per_page)
    daily_page = daily_paginator.get_page(request.GET.get('page'))

    punches_qs = qs.order_by(*PUNCH_LIST_ORDERING)
    punches_per_page = int(request.GET.get('punches_per_page') or 100)
    punches_paginator = Paginator(punches_qs, per_page=punches_per_page)
    punches_page = punches_paginator.get_page(request.GET.get('punches_page'))

    branches_qs = Branch.objects.filter(is_deleted=False, is_active=True).order_by('name')
    branch_ids = _user_accessible_branch_ids(request.user)
    if branch_ids is not None:
        branches_qs = branches_qs.filter(pk__in=branch_ids)

    mapped_filter = 'all'
    if filters['mapped_only'] is True:
        mapped_filter = 'yes'
    elif filters['mapped_only'] is False:
        mapped_filter = 'no'

    filter_employee = None
    if filters['employee_id']:
        filter_employee = Employee.objects.filter(
            pk=filters['employee_id'],
            is_deleted=False,
        ).first()

    return render(request, 'pages/attendance/report.html', {
        'daily_page': daily_page,
        'punches_page': punches_page,
        'summary': summary,
        'punch_stats': punch_stats,
        'total_daily_rows': len(all_rows),
        'devices': get_biometric_devices_queryset(request.user),
        'branches': branches_qs,
        'employees': Employee.objects.filter(is_deleted=False, status=Employee.Status.ACTIVE).order_by('name')[:500],
        'filter_employee': filter_employee,
        'filters': filters,
        'mapped_filter': mapped_filter,
        'querystring': _filters_to_querystring(filters),
        'daily_per_page': daily_per_page,
        'punches_per_page': punches_per_page,
        'punch_types': AttendancePunch.PunchType.choices,
    })


@permission_required('attendance.view')
@require_GET
def attendance_report_export(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة.')
        return redirect('web:attendance_report')

    filters = _apply_default_date_filters(_parse_filters(request))
    qs = _punches_for_report(request, filters)
    daily_table = daily_rows_to_table(build_daily_attendance_rows(qs))
    punch_table = punches_to_table_rows(qs.order_by(*PUNCH_LIST_ORDERING))

    wb = Workbook()
    header_fill = PatternFill('solid', fgColor='1E40AF')

    def _write_sheet(ws, table: dict, title: str) -> None:
        ws.title = title
        ws.sheet_view.rightToLeft = True
        for col, h in enumerate(table['columns'], 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = Font(bold=True, color='FFFFFF')
            c.alignment = Alignment(horizontal='center')
        for row_idx, row in enumerate(table['rows'], 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        for col in range(1, len(table['columns']) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

    ws_daily = wb.active
    _write_sheet(ws_daily, daily_table, 'يومي')
    ws_detail = wb.create_sheet('تفصيلي')
    _write_sheet(ws_detail, punch_table, 'تفصيلي')

    stamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    filename = f'attendance_report_{stamp}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
