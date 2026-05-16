"""سجلات الحضور — عرض تقني مع فلترة وتصفح وتصدير."""
from datetime import datetime, time

from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST

from apps.attendance.models import AttendancePunch, BiometricDevice
from apps.attendance.selectors.punch_records import (
    PUNCH_LIST_ORDERING,
    get_punch_queryset,
    get_punch_stats,
)
from apps.attendance.services.attendance_pull import pull_device_attendance
from apps.attendance.services.punch_inference import reclassify_punches_by_sequence
from apps.attendance.selectors.biometric_devices import (
    filter_biometric_devices_for_user,
    get_biometric_devices_queryset,
)
from apps.core.decorators import permission_required
from apps.core.models import Branch
from apps.core.web_views._helpers import _user_accessible_branch_ids
from apps.employees.models import Employee

FORCE_REAL_DEVICE = False


def _parse_filters(request) -> dict:
    branch_id = request.GET.get('branch') or None
    device_id = request.GET.get('device') or None
    employee_id = request.GET.get('employee') or None
    device_user_id = request.GET.get('device_user') or None
    date_from = request.GET.get('from') or None
    date_to = request.GET.get('to') or None
    punch_type = request.GET.get('punch_type') or None
    mapped = request.GET.get('mapped')
    mapped_only = None
    if mapped == '1':
        mapped_only = True
    elif mapped == '0':
        mapped_only = False
    return {
        'branch_id': int(branch_id) if branch_id and branch_id.isdigit() else None,
        'device_id': int(device_id) if device_id and device_id.isdigit() else None,
        'employee_id': int(employee_id) if employee_id and employee_id.isdigit() else None,
        'device_user_id': int(device_user_id) if device_user_id and device_user_id.isdigit() else None,
        'date_from': date_from,
        'date_to': date_to,
        'punch_type': punch_type if punch_type else None,
        'mapped_only': mapped_only,
        'search': (request.GET.get('q') or '').strip(),
    }


def _filters_to_querystring(filters: dict, *, extra: dict | None = None) -> str:
    from urllib.parse import urlencode

    params = {}
    if filters.get('branch_id'):
        params['branch'] = filters['branch_id']
    if filters.get('device_id'):
        params['device'] = filters['device_id']
    if filters.get('employee_id'):
        params['employee'] = filters['employee_id']
    if filters.get('device_user_id'):
        params['device_user'] = filters['device_user_id']
    if filters.get('date_from'):
        params['from'] = filters['date_from']
    if filters.get('date_to'):
        params['to'] = filters['date_to']
    if filters.get('punch_type'):
        params['punch_type'] = filters['punch_type']
    if filters.get('mapped_only') is True:
        params['mapped'] = '1'
    elif filters.get('mapped_only') is False:
        params['mapped'] = '0'
    if filters.get('search'):
        params['q'] = filters['search']
    if extra:
        params.update(extra)
    return urlencode(params)


@permission_required('attendance.view')
def attendance_records_list(request):
    filters = _parse_filters(request)
    date_from = None
    date_to = None
    if filters['date_from']:
        date_from = datetime.strptime(filters['date_from'], '%Y-%m-%d').date()
    if filters['date_to']:
        date_to = datetime.strptime(filters['date_to'], '%Y-%m-%d').date()

    qs = get_punch_queryset(
        device_id=filters['device_id'],
        branch_id=filters['branch_id'],
        employee_id=filters['employee_id'],
        device_user_id=filters['device_user_id'],
        date_from=date_from,
        date_to=date_to,
        punch_type=filters['punch_type'],
        mapped_only=filters['mapped_only'],
        search=filters['search'] or None,
    ).filter(device_id__in=filter_biometric_devices_for_user(request.user).values('pk'))
    stats = get_punch_stats(qs, device_id=filters['device_id'])
    qs = qs.order_by(*PUNCH_LIST_ORDERING)
    paginator = Paginator(qs, per_page=int(request.GET.get('per_page') or 100))
    page_obj = paginator.get_page(request.GET.get('page'))

    branches_qs = Branch.objects.filter(is_deleted=False, is_active=True).order_by('name')
    branch_ids = _user_accessible_branch_ids(request.user)
    if branch_ids is not None:
        branches_qs = branches_qs.filter(pk__in=branch_ids)
    devices = get_biometric_devices_queryset(
        request.user,
        branch_id=filters['branch_id'],
    )
    employees = Employee.objects.filter(is_deleted=False, status=Employee.Status.ACTIVE).order_by('name')[:300]

    mapped_filter = 'all'
    if filters['mapped_only'] is True:
        mapped_filter = 'yes'
    elif filters['mapped_only'] is False:
        mapped_filter = 'no'

    return render(request, 'pages/attendance/records.html', {
        'page_obj': page_obj,
        'stats': stats,
        'devices': devices,
        'branches': branches_qs,
        'employees': employees,
        'filters': filters,
        'mapped_filter': mapped_filter,
        'querystring': _filters_to_querystring(filters),
        'punch_types': AttendancePunch.PunchType.choices,
        'per_page': paginator.per_page,
    })


@permission_required('attendance.manage')
@require_POST
def attendance_records_pull(request):
    device_id = request.POST.get('device_id')
    if not device_id:
        messages.error(request, 'اختر جهازاً للسحب.')
        return redirect('web:attendance_records')

    from apps.attendance.selectors.biometric_devices import get_device_for_user

    device = get_device_for_user(request.user, int(device_id))
    if not device.branch_id:
        messages.error(request, f'حدّد فرعاً لجهاز «{device.name}» قبل السحب.')
        return redirect('web:attendance_records')
    date_from = request.POST.get('date_from') or None
    date_to = request.POST.get('date_to') or None
    df = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else None
    dt = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None

    result = pull_device_attendance(
        device,
        date_from=df,
        date_to=dt,
        import_db=True,
        force_mock=FORCE_REAL_DEVICE,
    )
    if result.ok:
        msg = (
            f'«{device.name}»: على الجهاز {result.punches_fetched} — '
            f'جديد {result.punches_new} — مستورد {result.imported} — '
            f'مكرر {result.skipped_duplicate}'
        )
        if result.skipped_time_filter:
            msg += f' — قديم {result.skipped_time_filter}'
        messages.success(request, msg)
    else:
        messages.error(request, result.error)
    return redirect('web:attendance_records')


@permission_required('attendance.manage')
@require_POST
def attendance_records_reclassify(request):
    device_id = request.POST.get('device_id')
    did = int(device_id) if device_id and str(device_id).isdigit() else None
    result = reclassify_punches_by_sequence(device_id=did, dry_run=False)
    messages.success(
        request,
        f'تم إعادة التصنيف بالتسلسل: دخول {result["inferred_in"]} · خروج {result["inferred_out"]} '
        f'(حدّث {result["updated"]} سجل)',
    )
    url = reverse('web:attendance_records')
    if did:
        url = f'{url}?device={did}'
    return redirect(url)


@permission_required('attendance.view')
@require_GET
def attendance_records_export(request):
    filters = _parse_filters(request)
    date_from = datetime.strptime(filters['date_from'], '%Y-%m-%d').date() if filters.get('date_from') else None
    date_to = datetime.strptime(filters['date_to'], '%Y-%m-%d').date() if filters.get('date_to') else None

    qs = get_punch_queryset(
        device_id=filters['device_id'],
        branch_id=filters['branch_id'],
        employee_id=filters['employee_id'],
        device_user_id=filters['device_user_id'],
        date_from=date_from,
        date_to=date_to,
        punch_type=filters['punch_type'],
        mapped_only=filters['mapped_only'],
        search=filters['search'] or None,
    ).order_by(*PUNCH_LIST_ORDERING)[:50000]

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'سجلات الحضور'
    ws.sheet_view.rightToLeft = True

    headers = [
        'معرف السجل ZK', 'التاريخ', 'الوقت', 'الجهاز', 'IP',
        'رقم المستخدم', 'الاسم على الجهاز', 'موظف HR', 'الرقم الوظيفي',
        'نوع الحركة', 'رمز الحركة', 'طريقة التحقق', 'رمز التحقق',
        'دفعة المزامنة',
    ]
    header_fill = PatternFill('solid', fgColor='1E40AF')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center')

    for row_idx, p in enumerate(qs, 2):
        local = timezone.localtime(p.punched_at)
        ws.append([
            p.device_record_uid or '',
            local.strftime('%Y-%m-%d'),
            local.strftime('%H:%M:%S'),
            p.device.name,
            p.device.ip_address,
            p.device_user_id,
            p.device_user_name or '',
            p.employee.name if p.employee else 'غير مربوط',
            p.employee.employee_number if p.employee else '',
            p.get_punch_type_display(),
            p.punch_type,
            p.verify_mode_label or '—',
            p.verify_mode if p.verify_mode is not None else '',
            p.sync_batch,
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    stamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
    filename = f'attendance_records_{stamp}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
