"""تصدير بيانات الموظف إلى Excel ملوّن (بيانات تبويب الموظف فقط)."""
import logging

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_POST
from django_ratelimit.decorators import ratelimit

from apps.employees.models import Employee
from apps.core.web_views._helpers import employee_branch_access_required
from apps.core.decorators import permission_required
from apps.core.salary_access import user_can_view_salary

logger = logging.getLogger(__name__)

def _employee_administration_label(employee):
    adm = getattr(employee, 'administration', None)
    if not adm:
        return None
    code = (getattr(adm, 'code', None) or '').strip()
    name = (getattr(adm, 'name', None) or '').strip()
    if code and name:
        return f'{code} — {name}'
    return code or name or None


@login_required
@permission_required('employees.view')
@employee_branch_access_required
def export_employee_salary_excel(request, employee_id):
    """يُنشئ ملف Excel ملوّن يحتوي على بيانات تبويب الموظف فقط."""
    from django.contrib import messages
    from django.shortcuts import redirect

    if not user_can_view_salary(request.user):
        messages.error(request, 'لا تملك صلاحية تصدير بيانات الرواتب.')
        return redirect('web:view_employee', employee_id=employee_id)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    employee = get_object_or_404(
        Employee.objects.select_related('administration'),
        id=employee_id,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "بيانات الموظف"
    ws.sheet_view.rightToLeft = True

    # ──────────── أنماط ────────────
    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    title_fill = PatternFill('solid', fgColor='1E40AF')

    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2563EB')

    label_font = Font(name='Arial', size=11, bold=True, color='1E293B')
    label_fill = PatternFill('solid', fgColor='F1F5F9')

    value_font = Font(name='Arial', size=11, color='0F172A')
    value_fill = PatternFill('solid', fgColor='FFFFFF')

    salary_label_fill = PatternFill('solid', fgColor='DBEAFE')
    salary_value_fill = PatternFill('solid', fgColor='EFF6FF')

    allowance_fill = PatternFill('solid', fgColor='ECFDF5')
    allowance_label_fill = PatternFill('solid', fgColor='D1FAE5')

    deduction_fill = PatternFill('solid', fgColor='FEF2F2')
    deduction_label_fill = PatternFill('solid', fgColor='FEE2E2')

    total_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    total_fill = PatternFill('solid', fgColor='059669')

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center', wrap_text=True)

    thin = Side(border_style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ──────────── بيانات الموظف كجدول أفقي (الأعمدة = الحقول) ────────────
    columns = [
        ("الاسم", employee.name),
        ("رقم الموظف", employee.employee_number),
        ("رقم الهوية", employee.id_number),
        ("الجوال", employee.phone),
        ("البريد الإلكتروني", employee.email),
        ("الجنسية", getattr(employee.nationality, 'name', None)),
        ("المهنة", getattr(employee.profession, 'name', None)),
        ("الكفالة", getattr(employee.sponsorship, 'name', None)),
        ("الفرع", getattr(employee.branch, 'name', None)),
        ("القسم", getattr(employee.department, 'name', None)),
        ("الإدارة", _employee_administration_label(employee)),
        ("مركز التكلفة", getattr(employee.cost_center, 'name', None)),
        ("تاريخ المباشرة", employee.hire_date),
        ("تاريخ التوقف", employee.end_date),
        ("تاريخ انتهاء التأمين الطبي", employee.medical_insurance_expiry_date),
        ("تاريخ انتهاء العقد (مخطط)", employee.contract_expiry_date),
        ("التأمين", getattr(employee.insurance, 'name', None)),
        ("فئة التأمين", getattr(employee.insurance_class, 'name', None)),
        ("الحالة", employee.get_status_display()),
        ("سبب الانتهاء", employee.end_reason),
    ]

    # عرض الأعمدة (مضغوط)
    for idx, (label, _v) in enumerate(columns, start=1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = 10

    n = len(columns)
    last_col = get_column_letter(n)

    # العنوان الكبير
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(row=1, column=1, value="بيانات الموظف")
    c.font = title_font
    c.fill = title_fill
    c.alignment = center
    c.border = border
    ws.row_dimensions[1].height = 38

    # تاريخ الإصدار
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    c = ws.cell(row=2, column=1, value=f"تاريخ الإصدار: {timezone.now().strftime('%Y-%m-%d %H:%M')}")
    c.font = Font(name='Arial', size=10, italic=True, color='64748B')
    c.alignment = center
    ws.row_dimensions[2].height = 22

    # صف العناوين (الأعمدة) - بخط أصغر والتفاف للأسطر
    header_row = 4
    compact_header_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    for idx, (label, _v) in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=idx, value=label)
        c.font = compact_header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[header_row].height = 60

    # صف القيم
    value_row = header_row + 1
    compact_value_font = Font(name='Arial', size=9, color='0F172A')
    for idx, (_label, value) in enumerate(columns, start=1):
        c = ws.cell(row=value_row, column=idx,
                    value=value if value not in (None, '') else '—')
        c.font = compact_value_font
        c.fill = value_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[value_row].height = 22

    # تجميد صف العناوين
    ws.freeze_panes = ws.cell(row=value_row, column=1)

    # ──────────── إخراج الملف ────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = (employee.name or 'employee').replace(' ', '_')
    filename = f"employee_{safe_name}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


_NO_SPONSORSHIP_COLUMNS = (
    ('الاسم', 'name'),
    ('رقم الهوية', 'id_number'),
    ('رقم الجوال', 'phone'),
    ('الرقم الوظيفي', 'employee_number'),
    ('تاريخ المباشرة', 'hire_date'),
    ('الجنسية', 'nationality'),
    ('المهنة', 'profession'),
    ('الراتب الأساسي', 'basic_salary'),
    ('الفرع', 'branch'),
    ('القسم', 'department'),
    ('الإدارة', 'administration'),
    ('مركز التكلفة', 'cost_center'),
)

_SALARY_COLUMN_KEYS = frozenset({'basic_salary'})


def _text_or_dash(value) -> str:
    text = (str(value).strip() if value is not None else '')
    return text or '—'


def _fk_name(obj) -> str:
    """اسم FK كما في تبويب بيانات الموظف (nationality.name / department.name / …)."""
    if obj is None:
        return '—'
    return _text_or_dash(getattr(obj, 'name', None))


def _no_sponsorship_cell_value(employee: Employee, key: str):
    """
    ربط أعمدة Excel بحقول Employee كما تُعرض في الموقع
    (pages/employees/_tab_main.html + تقارير الموظفين).
    """
    if key == 'name':
        return _text_or_dash(employee.name)
    if key == 'id_number':
        return _text_or_dash(employee.id_number)
    if key == 'phone':
        return _text_or_dash(employee.phone)
    if key == 'employee_number':
        return _text_or_dash(employee.employee_number)
    if key == 'hire_date':
        # تاريخ مباشرة الموظف — يُكتب كتاريخ Excel صحيح
        return employee.hire_date if employee.hire_date else '—'
    if key == 'nationality':
        return _fk_name(employee.nationality)
    if key == 'profession':
        return _fk_name(employee.profession)
    if key == 'basic_salary':
        # حقل الراتب الأساسي في تبويب الراتب
        value = employee.basic_salary
        if value is None:
            return '—'
        return float(value)
    if key == 'branch':
        return _fk_name(employee.branch)
    if key == 'department':
        return _fk_name(employee.department)
    if key == 'administration':
        # نفس عرض الملف: «رمز — اسم»
        return _employee_administration_label(employee) or '—'
    if key == 'cost_center':
        return _fk_name(employee.cost_center)
    return '—'


MAX_EXPORT_ROWS = 10000


def _excel_lock(user_id: int, kind: str, timeout: int) -> bool:
    """قفل قصير يمنع تكرار التصدير/الاستيراد لنفس المستخدم."""
    from django.core.cache import cache

    return bool(cache.add(f'employees:ns_excel:{kind}:{user_id}', '1', timeout=timeout))


def _excel_unlock(user_id: int, kind: str) -> None:
    from django.core.cache import cache

    cache.delete(f'employees:ns_excel:{kind}:{user_id}')


def _secure_xlsx_response(payload: bytes, filename: str) -> HttpResponse:
    response = HttpResponse(
        payload,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
    response['Pragma'] = 'no-cache'
    response['X-Content-Type-Options'] = 'nosniff'
    response['Content-Length'] = str(len(payload))
    return response


@login_required
@permission_required('employees.view')
@ratelimit(key='user', rate='20/h', method='GET', block=True, group='ns_emp_export')
def export_non_sponsored_employees_excel(request):
    """تصدير موظفين بدون كفالة — أعمدة مربوطة بحقول بيانات الموظف في الموقع."""
    from io import BytesIO

    from django.contrib import messages
    from django.shortcuts import redirect
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from apps.core.selectors.employee_search import apply_employee_search
    from apps.core.web_views._helpers import filter_employees_queryset_for_user

    if not _excel_lock(request.user.pk, 'export', timeout=90):
        messages.warning(request, 'تصدير قيد التنفيذ بالفعل. انتظر اكتماله ثم أعد المحاولة.')
        return redirect('web:list_employees')

    try:
        can_salary = user_can_view_salary(request.user)
        columns = [
            (label, key)
            for label, key in _NO_SPONSORSHIP_COLUMNS
            if can_salary or key not in _SALARY_COLUMN_KEYS
        ]

        qs = (
            Employee.objects.filter(
                is_deleted=False,
                sponsorship__isnull=True,
            )
            .select_related(
                'branch',
                'department',
                'administration',
                'cost_center',
                'nationality',
                'profession',
            )
            .only(
                'id',
                'name',
                'id_number',
                'phone',
                'employee_number',
                'hire_date',
                'basic_salary',
                'branch__name',
                'department__name',
                'administration__code',
                'administration__name',
                'cost_center__name',
                'nationality__name',
                'profession__name',
            )
            .order_by('name', 'id')
        )
        qs = filter_employees_queryset_for_user(request.user, qs)

        q = (request.GET.get('q') or '').strip()
        if q:
            qs = apply_employee_search(qs, q)

        # write_only يقلّل استهلاك الذاكرة؛ نتجنّب تنسيق كل خلية في الجسم
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title='بدون كفالة')
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass

        header_fill = PatternFill('solid', fgColor='1E40AF')
        header_font = Font(name='Arial', bold=True, color='FFFFFF')
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        header_cells = []
        for col, (label, _key) in enumerate(columns, 1):
            cell = WriteOnlyCell(ws, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            header_cells.append(cell)
            try:
                ws.column_dimensions[get_column_letter(col)].width = 18
            except Exception:
                pass
        ws.append(header_cells)

        exported = 0
        for employee in qs.iterator(chunk_size=500):
            if exported >= MAX_EXPORT_ROWS:
                break
            ws.append([_no_sponsorship_cell_value(employee, key) for _label, key in columns])
            exported += 1

        buf = BytesIO()
        wb.save(buf)
        stamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        filename = f'employees_no_sponsorship_{stamp}.xlsx'
        response = _secure_xlsx_response(buf.getvalue(), filename)
        if exported >= MAX_EXPORT_ROWS:
            response['X-HR-Export-Truncated'] = '1'
        return response
    except Exception:
        logger.exception('non-sponsored employees export failed user=%s', request.user.pk)
        messages.error(request, 'تعذّر إنشاء ملف التصدير. حاول مرة أخرى.')
        return redirect('web:list_employees')
    finally:
        _excel_unlock(request.user.pk, 'export')


@login_required
@permission_required('employees.add')
@require_POST
@ratelimit(key='user', rate='10/h', method='POST', block=True, group='ns_emp_import')
def import_non_sponsored_employees_excel(request):
    """استيراد موظفين بدون كفالة من Excel (نفس أعمدة التصدير)."""
    from django.contrib import messages
    from django.shortcuts import redirect

    from apps.core.salary_access import user_can_edit_salary
    from apps.core.services.access_control import get_accessible_branch_ids
    from apps.employees.services.non_sponsored_import import (
        ImportValidationError,
        import_non_sponsored_employees_from_upload,
    )

    if not _excel_lock(request.user.pk, 'import', timeout=180):
        messages.warning(
            request,
            'استيراد قيد التنفيذ بالفعل. انتظر حتى ينتهي لتفادي تكرار البيانات.',
        )
        return redirect('web:list_employees')

    try:
        uploaded = request.FILES.get('excel_file')
        if not uploaded:
            messages.error(request, 'يرجى اختيار ملف Excel للاستيراد.')
            return redirect('web:list_employees')

        allowed = get_accessible_branch_ids(request.user)
        allowed_set = set(allowed) if allowed is not None else None

        try:
            summary = import_non_sponsored_employees_from_upload(
                uploaded,
                user=request.user,
                allowed_branch_ids=allowed_set,
                apply_salary=user_can_edit_salary(request.user),
            )
        except ImportValidationError as exc:
            messages.error(request, str(exc))
            return redirect('web:list_employees')
        except Exception:
            logger.exception('non-sponsored employees import failed user=%s', request.user.pk)
            messages.error(request, 'تعذّر استيراد الملف. تأكد من صحة Excel وأعد المحاولة.')
            return redirect('web:list_employees')

        if summary.created or summary.updated:
            messages.success(
                request,
                f'تم الاستيراد: إنشاء {summary.created}، تحديث {summary.updated}.',
            )
        if summary.skipped:
            messages.warning(request, f'تم تخطي {summary.skipped} صفاً.')
        if summary.errors:
            detail_errors = [
                f'صف {r.row_number}: {r.message}'
                for r in summary.results
                if r.action == 'error'
            ][:5]
            extra = (' — ' + ' | '.join(detail_errors)) if detail_errors else ''
            messages.error(
                request,
                f'فشل استيراد {summary.errors} صفاً{extra}',
            )
        if not (summary.created or summary.updated or summary.skipped or summary.errors):
            messages.warning(request, 'لا توجد صفوف للاستيراد في الملف.')

        return redirect('web:list_employees')
    finally:
        _excel_unlock(request.user.pk, 'import')
