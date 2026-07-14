"""اختبارات تصدير موظفين بدون كفالة إلى Excel."""
from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from apps.core.models import Branch, Company
from apps.cost_centers.models import CostCenter
from apps.departments.models import Department
from apps.employees.models import Employee
from apps.setup.models import Administration, Nationality, Profession, Sponsorship

User = get_user_model()


@override_settings(ALLOWED_HOSTS=['testserver'])
class NonSponsoredEmployeesExcelExportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        company = Company.objects.create(name='شركة تصدير')
        cls.branch = Branch.objects.create(name='فرع حائل', code='EX1', company=company)
        cls.department = Department.objects.create(code='D1', name='المبيعات', branch=cls.branch)
        cls.administration = Administration.objects.create(code='A1', name='العمليات')
        cls.cost_center = CostCenter.objects.create(code='C1', name='مركز مبيعات', branch=cls.branch)
        cls.nationality = Nationality.objects.create(name='يمني', code='YE')
        cls.profession = Profession.objects.create(name='محاسب', code='ACC')
        cls.sponsorship = Sponsorship.objects.create(
            code='SP-EX',
            company_name='كفالة تصدير',
            is_active=True,
        )
        cls.admin = User.objects.create_user(
            username='export_admin',
            password='pass12345',
            is_superuser=True,
            is_staff=True,
        )
        cls.no_spons = Employee.objects.create(
            name='بدون كفالة',
            employee_number='9001',
            id_number='111222333',
            phone='0500000001',
            branch=cls.branch,
            department=cls.department,
            administration=cls.administration,
            cost_center=cls.cost_center,
            nationality=cls.nationality,
            profession=cls.profession,
            hire_date=date(2024, 1, 15),
            sponsorship=None,
            basic_salary=Decimal('3000.50'),
            housing_allowance=Decimal('500.00'),
            status=Employee.Status.ACTIVE,
        )
        cls.with_spons = Employee.objects.create(
            name='على كفالة',
            employee_number='9002',
            id_number='222',
            phone='0500000002',
            branch=cls.branch,
            sponsorship=cls.sponsorship,
            basic_salary=Decimal('4000.00'),
            housing_allowance=Decimal('700.00'),
            status=Employee.Status.ACTIVE,
        )

    def setUp(self):
        self.client = Client()
        self.client.force_login(self.admin)

    def test_export_includes_only_non_sponsored(self):
        response = self.client.get(reverse('web:export_non_sponsored_employees_excel'))
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response['Content-Type'],
        )
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertIn('الاسم', headers)
        self.assertIn('الراتب الأساسي', headers)
        self.assertIn('القسم', headers)
        self.assertIn('الإدارة', headers)
        self.assertIn('مركز التكلفة', headers)
        self.assertNotIn('بدل سكن', headers)
        self.assertNotIn('البريد الإلكتروني', headers)
        self.assertNotIn('نسبة خصم التأمينات', headers)
        self.assertNotIn('الحالة', headers)
        self.assertNotIn('الكفالة', headers)
        self.assertNotIn('البنك', headers)

        names = {row[0].value for row in ws.iter_rows(min_row=2, max_col=1)}
        self.assertIn('بدون كفالة', names)
        self.assertNotIn('على كفالة', names)

    def test_export_row_maps_site_fields(self):
        response = self.client.get(reverse('web:export_non_sponsored_employees_excel'))
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        row = [cell.value for cell in ws[2]]
        by_header = dict(zip(headers, row))

        self.assertEqual(by_header['الاسم'], 'بدون كفالة')
        self.assertEqual(by_header['رقم الهوية'], '111222333')
        self.assertEqual(by_header['رقم الجوال'], '0500000001')
        self.assertEqual(by_header['الرقم الوظيفي'], '9001')
        self.assertEqual(by_header['الجنسية'], 'يمني')
        self.assertEqual(by_header['المهنة'], 'محاسب')
        self.assertEqual(by_header['الفرع'], 'فرع حائل')
        self.assertEqual(by_header['القسم'], 'المبيعات')
        self.assertEqual(by_header['الإدارة'], 'A1 — العمليات')
        self.assertEqual(by_header['مركز التكلفة'], 'مركز مبيعات')
        self.assertEqual(float(by_header['الراتب الأساسي']), 3000.50)
        self.assertTrue(str(by_header['تاريخ المباشرة']).startswith('2024-01-15'))

    def test_export_requires_login(self):
        self.client.logout()
        response = self.client.get(reverse('web:export_non_sponsored_employees_excel'))
        self.assertEqual(response.status_code, 302)

    def test_list_page_has_export_button(self):
        response = self.client.get(reverse('web:list_employees'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Excel بدون كفالة')
        self.assertContains(response, reverse('web:export_non_sponsored_employees_excel'))

    def test_export_sets_secure_headers(self):
        response = self.client.get(reverse('web:export_non_sponsored_employees_excel'))
        self.assertEqual(response.status_code, 200)
        self.assertIn('no-store', response['Cache-Control'])
        self.assertEqual(response['X-Content-Type-Options'], 'nosniff')
