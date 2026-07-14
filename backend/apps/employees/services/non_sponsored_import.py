"""استيراد موظفين بدون كفالة من Excel بنفس أعمدة التصدير.

مصمَّم للأمان والأداء:
- حد حجم الملف وعدد الصفوف
- التحقق من أن الملف أرشيف xlsx حقيقي
- كاش كامل للمراجع (فروع، أقسام، …) بدل استعلام لكل صف
- جلب الموظفين الموجودين دفعة واحدة
- حفظ على دفعات atomic لتقليل قفل قاعدة البيانات
"""
from __future__ import annotations

import io
import logging
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from django.db import transaction
from django.db.models import Q
from django.utils import timezone

from apps.core.models import Branch
from apps.cost_centers.models import CostCenter
from apps.departments.models import Department
from apps.employees.models import Employee
from apps.setup.models import Nationality, Profession

logger = logging.getLogger(__name__)

# يطابق `_NO_SPONSORSHIP_COLUMNS` في employee_export.py
HEADER_TO_FIELD = {
    'الاسم': 'name',
    'رقم الهوية': 'id_number',
    'رقم الجوال': 'phone',
    'الجوال': 'phone',
    'الرقم الوظيفي': 'employee_number',
    'رقم الموظف': 'employee_number',
    'تاريخ المباشرة': 'hire_date',
    'الجنسية': 'nationality',
    'المهنة': 'profession',
    'الراتب الأساسي': 'basic_salary',
    'الفرع': 'branch',
    'القسم': 'department',
    'مركز التكلفة': 'cost_center',
}

MAX_IMPORT_FILE_BYTES = 5 * 1024 * 1024  # 5MB
MAX_IMPORT_ROWS = 3000
IMPORT_SAVE_BATCH_SIZE = 100
MAX_DETAIL_RESULTS = 40

_FIELD_MAX_LEN = {
    'name': 200,
    'id_number': 50,
    'phone': 20,
    'employee_number': 50,
}


@dataclass
class ImportRowResult:
    row_number: int
    action: str  # created | updated | skipped | error
    message: str
    employee_number: str = ''


@dataclass
class ImportSummary:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0
    results: list[ImportRowResult] = field(default_factory=list)

    def add_result(self, result: ImportRowResult) -> None:
        if result.action == 'created':
            self.created += 1
        elif result.action == 'updated':
            self.updated += 1
        elif result.action == 'skipped':
            self.skipped += 1
        elif result.action == 'error':
            self.errors += 1
        if len(self.results) < MAX_DETAIL_RESULTS or result.action == 'error':
            if len(self.results) < MAX_DETAIL_RESULTS:
                self.results.append(result)


class ImportValidationError(ValueError):
    """خطأ تحقق آمن للعرض للمستخدم."""


def validate_excel_upload(uploaded_file) -> None:
    """يفحص الامتداد والحجم وأن الملف أرشيف xlsx صالح."""
    name = (getattr(uploaded_file, 'name', '') or '').lower()
    if not name.endswith('.xlsx'):
        raise ImportValidationError('يُقبل فقط ملف Excel بصيغة .xlsx')

    size = getattr(uploaded_file, 'size', None)
    if size is not None and size > MAX_IMPORT_FILE_BYTES:
        raise ImportValidationError('حجم الملف يتجاوز الحد المسموح (5 ميغابايت).')
    if size == 0:
        raise ImportValidationError('الملف فارغ.')

    position = uploaded_file.tell() if hasattr(uploaded_file, 'tell') else None
    try:
        if hasattr(uploaded_file, 'seek'):
            uploaded_file.seek(0)
        header = uploaded_file.read(4)
        if not isinstance(header, bytes):
            header = bytes(header or b'')
        if not header.startswith(b'PK\x03\x04'):
            raise ImportValidationError('محتوى الملف ليس Excel صالحاً.')

        if hasattr(uploaded_file, 'seek'):
            uploaded_file.seek(0)
        # حد قراءة للتحقق من البنية دون تحميل ضخم في الذاكرة إن أمكن
        data = uploaded_file.read(MAX_IMPORT_FILE_BYTES + 1)
        if not isinstance(data, bytes):
            data = bytes(data or b'')
        if len(data) > MAX_IMPORT_FILE_BYTES:
            raise ImportValidationError('حجم الملف يتجاوز الحد المسموح (5 ميغابايت).')
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as archive:
                names = archive.namelist()
                if not any(n.startswith('xl/') for n in names):
                    raise ImportValidationError('محتوى الملف لا يطابق صيغة Excel.')
                # حماية بسيطة من أرشيف ضخم جداً (zip bomb)
                total_uncompressed = 0
                for info in archive.infolist():
                    total_uncompressed += max(info.file_size, 0)
                    if total_uncompressed > 40 * 1024 * 1024:
                        raise ImportValidationError('ملف Excel غير آمن أو كبير جداً بعد فك الضغط.')
        except zipfile.BadZipFile as exc:
            raise ImportValidationError('محتوى الملف ليس Excel صالحاً.') from exc
    finally:
        if hasattr(uploaded_file, 'seek'):
            uploaded_file.seek(0 if position is None else position)


def _norm_header(value: Any) -> str:
    if value is None:
        return ''
    return str(value).strip().replace('\n', ' ')


def _sanitize_text(value: str, max_len: int) -> str:
    text = (value or '').strip()
    if text and text[0] in ('=', '+', '-', '@'):
        text = "'" + text
    return text[:max_len]


def _cell_str(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text in {'—', '-', '–', 'N/A', 'n/a'}:
        return ''
    return text


def _cell_date(value: Any) -> date | None:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _cell_str(value)
    if not text:
        return None
    text = text[:10]
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ImportValidationError(f'تاريخ غير صالح: {value!r}')


def _cell_decimal(value: Any) -> Decimal | None:
    if value is None or value == '':
        return None
    text = _cell_str(value).replace(',', '')
    if not text:
        return None
    try:
        amount = Decimal(text).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError) as exc:
        raise ImportValidationError(f'رقم غير صالح: {value!r}') from exc
    if amount < 0 or amount > Decimal('9999999999.99'):
        raise ImportValidationError('قيمة الراتب خارج النطاق المسموح.')
    return amount


def _parse_header_map(header_row) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        label = _norm_header(cell)
        field = HEADER_TO_FIELD.get(label)
        if field and field not in mapping:
            mapping[field] = idx
    return mapping


def _row_dict(row, header_map: dict[str, int]) -> dict[str, Any]:
    data: dict[str, Any] = {}
    for field, idx in header_map.items():
        data[field] = row[idx] if idx < len(row) else None
    return data


def _index_lookup(objects) -> dict[str, Any]:
    """فهرس بالاسم والرمز (صغير الحروف) للبحث O(1)."""
    index: dict[str, Any] = {}
    for obj in objects:
        name = (getattr(obj, 'name', None) or '').strip().lower()
        if name and name not in index:
            index[name] = obj
        code = (getattr(obj, 'code', None) or '').strip().lower()
        if code and code not in index:
            index[code] = obj
    return index


def _resolve_from_index(index: dict[str, Any], raw: str):
    key = (raw or '').strip().lower()
    if not key:
        return None
    return index.get(key)


class _ReferenceCache:
    def __init__(self, allowed_branch_ids: set[int] | None):
        self.allowed_branch_ids = allowed_branch_ids
        branch_qs = Branch.objects.filter(is_deleted=False, is_active=True)
        if allowed_branch_ids is not None:
            branch_qs = branch_qs.filter(pk__in=allowed_branch_ids)
        self.branches = _index_lookup(branch_qs)
        self.nationalities = _index_lookup(
            Nationality.objects.filter(is_deleted=False, is_active=True),
        )
        self.professions = _index_lookup(
            Profession.objects.filter(is_deleted=False, is_active=True),
        )
        self.departments = _index_lookup(
            Department.objects.filter(is_deleted=False, is_active=True),
        )
        self.cost_centers = _index_lookup(
            CostCenter.objects.filter(is_deleted=False, is_active=True),
        )


def _load_existing_maps(
    employee_numbers: set[str],
    id_numbers: set[str],
) -> tuple[dict[str, Employee], dict[str, Employee]]:
    by_number: dict[str, Employee] = {}
    by_id: dict[str, Employee] = {}
    if not employee_numbers and not id_numbers:
        return by_number, by_id

    q = Q()
    if employee_numbers:
        q |= Q(employee_number__in=employee_numbers)
    if id_numbers:
        q |= Q(id_number__in=id_numbers)

    for emp in Employee.objects.filter(is_deleted=False).filter(q).only(
        'id',
        'name',
        'employee_number',
        'id_number',
        'phone',
        'hire_date',
        'basic_salary',
        'branch_id',
        'department_id',
        'cost_center_id',
        'nationality_id',
        'profession_id',
        'sponsorship_id',
        'status',
    ):
        if emp.employee_number:
            by_number.setdefault(emp.employee_number, emp)
        if emp.id_number:
            by_id.setdefault(emp.id_number, emp)
    return by_number, by_id


def _find_existing_cached(
    employee_number: str,
    id_number: str,
    *,
    by_number: dict[str, Employee],
    by_id: dict[str, Employee],
    allowed_branch_ids: set[int] | None,
) -> Employee | None:
    emp = None
    if employee_number:
        emp = by_number.get(employee_number)
    if emp is None and id_number:
        emp = by_id.get(id_number)
    if emp is None:
        return None
    if allowed_branch_ids is not None and emp.branch_id not in allowed_branch_ids:
        raise ImportValidationError(
            f'الموظف موجود في فرع خارج صلاحيتك ({emp.employee_number or emp.id_number})',
        )
    return emp


def _apply_row_to_employee(
    emp: Employee,
    *,
    name: str,
    employee_number: str,
    id_number: str,
    phone: str,
    hire_date: date | None,
    branch,
    nationality,
    profession,
    department,
    cost_center,
    basic_salary: Decimal | None,
    is_create: bool,
    allowed_branch_ids: set[int] | None,
) -> None:
    emp.name = name
    emp.sponsorship = None
    if employee_number:
        emp.employee_number = employee_number
    if id_number:
        emp.id_number = id_number
    if phone:
        emp.phone = phone
    if hire_date is not None:
        emp.hire_date = hire_date
    if branch is not None:
        emp.branch = branch
    elif is_create and allowed_branch_ids is not None and len(allowed_branch_ids) == 1:
        emp.branch_id = next(iter(allowed_branch_ids))
    if nationality is not None:
        emp.nationality = nationality
    if profession is not None:
        emp.profession = profession
    if department is not None:
        emp.department = department
    if cost_center is not None:
        emp.cost_center = cost_center
    if basic_salary is not None:
        emp.basic_salary = basic_salary
    if is_create and emp.branch_id is None and allowed_branch_ids is not None:
        raise ImportValidationError('الفرع مطلوب عند إنشاء موظف جديد')
    emp.updated_at = timezone.now()


def import_non_sponsored_employees_from_workbook(
    workbook,
    *,
    user,
    allowed_branch_ids: set[int] | None,
    apply_salary: bool = True,
) -> ImportSummary:
    """
    يستورد صفوف الشيت الأول بنفس أعمدة التصدير.
    - ينشئ موظفاً جديداً أو يحدّث موجوداً (بدون كفالة فقط).
    - يتخطى موظفاً مسجّلاً على كفالة.
    """
    del user  # reserved for audit hooks
    summary = ImportSummary()
    ws = workbook.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        summary.add_result(ImportRowResult(1, 'error', 'الملف فارغ'))
        return summary

    header_map = _parse_header_map(header_row)
    if 'name' not in header_map:
        summary.add_result(ImportRowResult(1, 'error', 'عمود «الاسم» مطلوب في الصف الأول'))
        return summary

    raw_rows: list[tuple[int, tuple]] = []
    for excel_row_num, row in enumerate(rows_iter, start=2):
        if row is None or all(c is None or str(c).strip() == '' for c in row):
            continue
        raw_rows.append((excel_row_num, row))
        if len(raw_rows) > MAX_IMPORT_ROWS:
            raise ImportValidationError(
                f'عدد الصفوف يتجاوز الحد المسموح ({MAX_IMPORT_ROWS}). قسّم الملف إلى أجزاء.',
            )

    refs = _ReferenceCache(allowed_branch_ids)

    employee_numbers: set[str] = set()
    id_numbers: set[str] = set()
    for _num, row in raw_rows:
        raw = _row_dict(row, header_map)
        en = _sanitize_text(_cell_str(raw.get('employee_number')), _FIELD_MAX_LEN['employee_number'])
        idn = _sanitize_text(_cell_str(raw.get('id_number')), _FIELD_MAX_LEN['id_number'])
        if en:
            employee_numbers.add(en)
        if idn:
            id_numbers.add(idn)

    by_number, by_id = _load_existing_maps(employee_numbers, id_numbers)

    pending_creates: list[tuple[Employee, ImportRowResult]] = []
    pending_updates: list[tuple[Employee, ImportRowResult]] = []

    def flush_pending() -> None:
        nonlocal pending_creates, pending_updates
        if not pending_creates and not pending_updates:
            return
        with transaction.atomic():
            for emp, result in pending_creates:
                emp.save()
                if emp.employee_number:
                    by_number[emp.employee_number] = emp
                if emp.id_number:
                    by_id[emp.id_number] = emp
                summary.add_result(result)
            for emp, result in pending_updates:
                emp.save(update_fields=[
                    'name', 'sponsorship', 'employee_number', 'id_number', 'phone',
                    'hire_date', 'branch', 'nationality', 'profession', 'department',
                    'cost_center', 'basic_salary', 'updated_at',
                ])
                if emp.employee_number:
                    by_number[emp.employee_number] = emp
                if emp.id_number:
                    by_id[emp.id_number] = emp
                summary.add_result(result)
        pending_creates = []
        pending_updates = []

    for excel_row_num, row in raw_rows:
        raw = _row_dict(row, header_map)
        name = _sanitize_text(_cell_str(raw.get('name')), _FIELD_MAX_LEN['name'])
        if not name:
            summary.add_result(
                ImportRowResult(excel_row_num, 'skipped', 'اسم فارغ — تم التخطي'),
            )
            continue

        try:
            employee_number = _sanitize_text(
                _cell_str(raw.get('employee_number')),
                _FIELD_MAX_LEN['employee_number'],
            )
            id_number = _sanitize_text(
                _cell_str(raw.get('id_number')),
                _FIELD_MAX_LEN['id_number'],
            )
            phone = _sanitize_text(_cell_str(raw.get('phone')), _FIELD_MAX_LEN['phone'])
            hire_date = _cell_date(raw.get('hire_date')) if 'hire_date' in header_map else None
            basic_salary = (
                _cell_decimal(raw.get('basic_salary'))
                if apply_salary and 'basic_salary' in header_map
                else None
            )

            branch = None
            if 'branch' in header_map:
                branch_raw = _cell_str(raw.get('branch'))
                if branch_raw:
                    branch = _resolve_from_index(refs.branches, branch_raw)
                    if branch is None:
                        raise ImportValidationError(
                            f'الفرع غير موجود أو خارج صلاحيتك: {branch_raw}',
                        )

            nationality = None
            if 'nationality' in header_map:
                nat_raw = _cell_str(raw.get('nationality'))
                if nat_raw:
                    nationality = _resolve_from_index(refs.nationalities, nat_raw)
                    if nationality is None:
                        raise ImportValidationError(f'الجنسية غير موجودة: {nat_raw}')

            profession = None
            if 'profession' in header_map:
                prof_raw = _cell_str(raw.get('profession'))
                if prof_raw:
                    profession = _resolve_from_index(refs.professions, prof_raw)
                    if profession is None:
                        raise ImportValidationError(f'المهنة غير موجودة: {prof_raw}')

            department = None
            if 'department' in header_map:
                dept_raw = _cell_str(raw.get('department'))
                if dept_raw:
                    department = _resolve_from_index(refs.departments, dept_raw)
                    if department is None:
                        raise ImportValidationError(f'القسم غير موجود: {dept_raw}')

            cost_center = None
            if 'cost_center' in header_map:
                cc_raw = _cell_str(raw.get('cost_center'))
                if cc_raw:
                    cost_center = _resolve_from_index(refs.cost_centers, cc_raw)
                    if cost_center is None:
                        raise ImportValidationError(f'مركز التكلفة غير موجود: {cc_raw}')

            existing = _find_existing_cached(
                employee_number,
                id_number,
                by_number=by_number,
                by_id=by_id,
                allowed_branch_ids=allowed_branch_ids,
            )
            if existing and existing.sponsorship_id:
                summary.add_result(
                    ImportRowResult(
                        excel_row_num,
                        'skipped',
                        f'الموظف «{existing.name}» على كفالة — لم يُعدَّل',
                        employee_number=existing.employee_number or '',
                    ),
                )
                continue

            if existing:
                emp = existing
                action = 'updated'
                is_create = False
            else:
                emp = Employee(status=Employee.Status.ACTIVE)
                action = 'created'
                is_create = True

            _apply_row_to_employee(
                emp,
                name=name,
                employee_number=employee_number,
                id_number=id_number,
                phone=phone,
                hire_date=hire_date,
                branch=branch,
                nationality=nationality,
                profession=profession,
                department=department,
                cost_center=cost_center,
                basic_salary=basic_salary,
                is_create=is_create,
                allowed_branch_ids=allowed_branch_ids,
            )

            result = ImportRowResult(
                excel_row_num,
                action,
                f'تم {"الإنشاء" if is_create else "التحديث"}: {emp.name}',
                employee_number=emp.employee_number or '',
            )
            if is_create:
                pending_creates.append((emp, result))
            else:
                pending_updates.append((emp, result))

            if len(pending_creates) + len(pending_updates) >= IMPORT_SAVE_BATCH_SIZE:
                flush_pending()

        except ImportValidationError as exc:
            summary.add_result(
                ImportRowResult(
                    excel_row_num,
                    'error',
                    str(exc),
                    employee_number=_sanitize_text(
                        _cell_str(raw.get('employee_number')),
                        _FIELD_MAX_LEN['employee_number'],
                    ),
                ),
            )
        except Exception:
            logger.exception('non-sponsored import row %s failed', excel_row_num)
            summary.add_result(
                ImportRowResult(
                    excel_row_num,
                    'error',
                    'تعذّر حفظ الصف لأسباب داخلية.',
                    employee_number=_sanitize_text(
                        _cell_str(raw.get('employee_number')),
                        _FIELD_MAX_LEN['employee_number'],
                    ),
                ),
            )

    flush_pending()
    return summary


def import_non_sponsored_employees_from_upload(
    uploaded_file,
    *,
    user,
    allowed_branch_ids,
    apply_salary=True,
) -> ImportSummary:
    from openpyxl import load_workbook

    validate_excel_upload(uploaded_file)
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    try:
        return import_non_sponsored_employees_from_workbook(
            wb,
            user=user,
            allowed_branch_ids=allowed_branch_ids,
            apply_salary=apply_salary,
        )
    finally:
        wb.close()
