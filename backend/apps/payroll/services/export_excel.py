"""تصدير مسير الرواتب إلى Excel منظم واحترافي."""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from django.utils import timezone

from apps.payroll.services.payroll_line_columns import (
    GROUP_HEADER_LABELS,
    PAYROLL_LINE_COLUMNS,
    payroll_lines_select_related,
    resolve_cell_value,
)

PAYROLL_EXPORT_COLUMNS = PAYROLL_LINE_COLUMNS

GROUP_HEADER_COLORS = {
    'info': '312E81',
    'earning': '047857',
    'deduction': 'B91C1C',
    'total': '1D4ED8',
}

GROUP_DATA_FILLS = {
    'info': 'F8FAFC',
    'earning': 'F0FDF4',
    'deduction': 'FFF5F5',
    'total': 'EFF6FF',
}

COLUMN_WIDTHS = {
    'employee_number': 11,
    'employee': 22,
    'branch': 14,
    'department': 14,
    'nationality': 12,
    'id_number': 14,
    'salary_period': 13,
    'salary_mode': 11,
    'bank': 16,
    'iban': 26,
    'basic_salary': 12,
    'insurance_deduction': 11,
    'other_deduction': 12,
    'total_earnings': 14,
    'total_deductions': 14,
    'net_salary': 13,
}

ROW_HEIGHT_TABLE = 20


def _money(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, Decimal):
        return float(val)
    return float(val)


def _style_merged_row(ws, row: int, col_start: int, col_end: int, *, font, fill, alignment, border):
    """تنسيق صف مدمج مع حدود وخلفية لكل الخلايا."""
    if col_end > col_start:
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row, end_column=col_end,
        )
    anchor = ws.cell(row=row, column=col_start)
    anchor.font = font
    anchor.fill = fill
    anchor.alignment = alignment
    anchor.border = border
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.border = border
        if col != col_start:
            cell.alignment = alignment


def build_payroll_run_workbook(run):
    """يُنشئ Workbook ملوّن ومنظم لمسير رواتب."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    ws.title = 'مسير الرواتب'[:31]
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

    n_cols = len(PAYROLL_EXPORT_COLUMNS)
    last_col_letter = get_column_letter(n_cols)

    thin = Side(border_style='thin', color='CBD5E1')
    medium = Side(border_style='medium', color='64748B')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_medium = Border(left=medium, right=medium, top=medium, bottom=medium)

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=False)
    align_right = Alignment(horizontal='right', vertical='center', wrap_text=False)
    align_left_ltr = Alignment(horizontal='left', vertical='center', wrap_text=False)

    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    meta_font = Font(name='Calibri', size=10, color='1E293B')
    meta_bold = Font(name='Calibri', size=10, bold=True, color='0F172A')
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10, color='0F172A')
    data_bold = Font(name='Calibri', size=10, bold=True, color='065F46')
    total_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')

    title_fill = PatternFill('solid', fgColor='1E3A8A')
    meta_fill = PatternFill('solid', fgColor='EEF2FF')
    footer_fill = PatternFill('solid', fgColor='0F766E')
    net_fill = PatternFill('solid', fgColor='D1FAE5')

    exported_at = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
    sponsorship = run.sponsorship.company_name if run.sponsorship_id and run.sponsorship else '—'

    # ── عنوان المسير ──
    _style_merged_row(
        ws, 1, 1, n_cols,
        font=title_font, fill=title_fill, alignment=align_center, border=border_medium,
    )
    ws.cell(row=1, column=1).value = f'مسير الرواتب — {run.branch.name} — {run.period_label}'
    ws.row_dimensions[1].height = 28

    meta_rows = [
        ('الفرع', run.branch.name),
        ('الفترة', run.period_label),
        ('نوع الراتب', run.get_salary_mode_display()),
        ('شركة الكفالة', sponsorship),
        ('الحالة', run.get_status_display()),
        ('عدد الموظفين', run.employees_count),
        ('تاريخ التصدير', exported_at),
    ]
    row = 2
    for label, value in meta_rows:
        ws.cell(row=row, column=1, value=label).font = meta_bold
        ws.cell(row=row, column=1).fill = meta_fill
        ws.cell(row=row, column=1).alignment = align_right
        ws.cell(row=row, column=1).border = border
        _style_merged_row(
            ws, row, 2, min(8, n_cols),
            font=meta_font, fill=meta_fill, alignment=align_right, border=border,
        )
        ws.cell(row=row, column=2, value=value)
        ws.row_dimensions[row].height = ROW_HEIGHT_TABLE
        row += 1

    summary_row = row + 1
    _style_merged_row(
        ws, summary_row, 1, n_cols,
        font=Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
        fill=PatternFill('solid', fgColor='059669'),
        alignment=align_center,
        border=border_medium,
    )
    ws.cell(row=summary_row, column=1).value = (
        f'إجمالي الاستحقاقات: {_money(run.total_earnings):,.2f}  |  '
        f'إجمالي الخصومات: {_money(run.total_deductions):,.2f}  |  '
        f'الصافي: {_money(run.total_net):,.2f}'
    )
    ws.row_dimensions[summary_row].height = ROW_HEIGHT_TABLE

    group_row = summary_row + 2
    header_row = group_row + 1
    data_start = header_row + 1

    # ── صف مجموعات الأعمدة ──
    col_idx = 1
    while col_idx <= n_cols:
        _k, _l, group, _t = PAYROLL_EXPORT_COLUMNS[col_idx - 1]
        start = col_idx
        while col_idx <= n_cols and PAYROLL_EXPORT_COLUMNS[col_idx - 1][2] == group:
            col_idx += 1
        end = col_idx - 1
        group_fill = PatternFill('solid', fgColor=GROUP_HEADER_COLORS.get(group, '475569'))
        _style_merged_row(
            ws, group_row, start, end,
            font=header_font, fill=group_fill, alignment=align_center, border=border_medium,
        )
        ws.cell(row=group_row, column=start, value=GROUP_HEADER_LABELS.get(group, group))
    ws.row_dimensions[group_row].height = ROW_HEIGHT_TABLE

    # ── رؤوس الأعمدة + عرض الأعمدة ──
    for idx, (key, label, group, col_type) in enumerate(PAYROLL_EXPORT_COLUMNS, start=1):
        header_fill = PatternFill('solid', fgColor=GROUP_HEADER_COLORS.get(group, '475569'))
        hc = ws.cell(row=header_row, column=idx, value=label)
        hc.font = header_font
        hc.fill = header_fill
        hc.alignment = align_center
        hc.border = border
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = COLUMN_WIDTHS.get(key, 12)

    ws.row_dimensions[header_row].height = ROW_HEIGHT_TABLE
    ws.freeze_panes = ws.cell(row=data_start, column=1).coordinate

    lines = list(payroll_lines_select_related(run.lines).order_by('employee__name'))
    sum_keys = {k for k, _l, _g, t in PAYROLL_EXPORT_COLUMNS if t in ('money', 'days')}
    totals = {key: 0.0 for key in sum_keys}

    # ── بيانات الموظفين (ارتفاع صف 20، محتوى داخل الخلية) ──
    for i, line in enumerate(lines):
        r = data_start + i
        row_fill = 'FFFFFF' if i % 2 == 0 else 'F1F5F9'
        for col_idx, (key, _label, group, col_type) in enumerate(PAYROLL_EXPORT_COLUMNS, start=1):
            raw = resolve_cell_value(line, run, key)
            if col_type == 'text':
                val = raw or '—'
            else:
                val = _money(raw)
                totals[key] += val

            if key == 'net_salary':
                fill = net_fill
                font = data_bold
            else:
                fill = PatternFill(
                    'solid',
                    fgColor=GROUP_DATA_FILLS.get(group, row_fill) if i % 2 == 0 else row_fill,
                )
                font = data_font

            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = font
            cell.fill = fill
            cell.border = border
            if key == 'iban':
                cell.alignment = align_left_ltr
            elif col_type == 'text':
                cell.alignment = (
                    align_right
                    if key in ('employee', 'employee_number', 'department', 'bank', 'branch')
                    else align_center
                )
            else:
                cell.alignment = align_center
            if col_type == 'money':
                cell.number_format = '#,##0.00'
            elif col_type == 'days':
                cell.number_format = '0.0'

        ws.row_dimensions[r].height = ROW_HEIGHT_TABLE

    # ── صف الإجمالي ──
    footer_row = data_start + len(lines)
    ws.row_dimensions[footer_row].height = ROW_HEIGHT_TABLE

    for col_idx, (key, _label, group, col_type) in enumerate(PAYROLL_EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=footer_row, column=col_idx)
        cell.font = total_font
        cell.fill = footer_fill
        cell.border = border_medium
        cell.alignment = align_center

        if key == 'employee':
            cell.value = 'الإجمالي'
            cell.alignment = align_right
        elif col_type == 'text':
            cell.value = None
        elif key in sum_keys:
            cell.value = totals[key]
            if col_type == 'money':
                cell.number_format = '#,##0.00'
            elif col_type == 'days':
                cell.number_format = '0.0'
        else:
            cell.value = None

    # ── جدول Excel رسمي (حدود، ترويسة، صفوف متناوبة) ──
    if lines:
        table_ref = f'A{header_row}:{last_col_letter}{footer_row}'
        table = Table(displayName='PayrollRunTable', ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name='TableStyleMedium2',
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    ws.print_options.horizontalCentered = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'

    return wb


def payroll_run_excel_filename(run) -> str:
    return f'payroll_{run.branch_id}_{run.period_year}_{run.period_month:02d}.xlsx'


def workbook_to_response(wb, filename: str):
    from django.http import HttpResponse

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
